"""
Base parser class for Excel file parsing
"""

from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class BaseExcelParser(ABC):
    """Abstract base class for Excel file parsing using Template Method pattern."""

    SHEET_NAME: str = ""  # Override in subclass
    METADATA_ROWS: int = 0  # Number of header/metadata rows to skip
    DATA_START_ROW: int = 1  # Row where data begins (1-indexed)

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet: Optional[Worksheet] = None

    def open_workbook(self) -> None:
        """Open the Excel workbook."""
        self.workbook = load_workbook(str(self.file_path), data_only=True, read_only=True)

    def close_workbook(self) -> None:
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()

    def get_sheet(self) -> Worksheet:
        """Get the target worksheet."""
        if not self.workbook:
            raise RuntimeError("Workbook not opened")

        # Try exact name first
        if self.SHEET_NAME and self.SHEET_NAME in self.workbook.sheetnames:
            return self.workbook[self.SHEET_NAME]

        # Try partial match
        for name in self.workbook.sheetnames:
            if self.SHEET_NAME and self.SHEET_NAME in name:
                return self.workbook[name]

        # Fall back to first sheet
        return self.workbook.active

    def get_cell_value(self, row: int, col: int) -> Any:
        """Get cell value by row and column (1-indexed)."""
        if self.worksheet is None:
            raise RuntimeError("Worksheet not set")
        return self.worksheet.cell(row=row, column=col).value

    def get_row_values(self, row: int, start_col: int = 1, end_col: Optional[int] = None) -> List[Any]:
        """Get all values from a row."""
        if self.worksheet is None:
            raise RuntimeError("Worksheet not set")

        if end_col is None:
            end_col = self.worksheet.max_column

        return [self.worksheet.cell(row=row, column=c).value for c in range(start_col, end_col + 1)]

    @abstractmethod
    def parse_filename(self) -> Dict[str, Any]:
        """Extract metadata from filename pattern. Override in subclass."""
        pass

    @abstractmethod
    def extract_metadata(self) -> Dict[str, Any]:
        """Extract header/metadata from worksheet. Override in subclass."""
        pass

    @abstractmethod
    def extract_data_rows(self) -> List[Dict[str, Any]]:
        """Extract data rows from worksheet. Override in subclass."""
        pass

    def run(self) -> Dict[str, Any]:
        """Template method orchestrating the parsing process."""
        try:
            self.open_workbook()
            self.worksheet = self.get_sheet()

            # Parse filename metadata
            file_meta = self.parse_filename()

            # Extract worksheet metadata
            sheet_meta = self.extract_metadata()

            # Merge metadata
            metadata = {**file_meta, **sheet_meta}

            # Extract data rows
            records = self.extract_data_rows()

            return {
                "filename": self.file_path.name,
                "metadata": metadata,
                "records": records
            }
        finally:
            self.close_workbook()
